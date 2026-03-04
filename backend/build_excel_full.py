from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json, sys

# ── Color Palette ─────────────────────────────────────────
BLUE   = '1D4ED8'; BLUE_L  = 'DBEAFE'
GREEN  = '10B981'; GREEN_L = 'D1FAE5'
RED    = 'EF4444'; RED_L   = 'FEE2E2'
PURPLE = '8B5CF6'; PURP_L  = 'EDE9FE'
AMBER  = 'F59E0B'; AMB_L   = 'FEF3C7'
TEAL   = '0891B2'; TEAL_L  = 'CCFBF1'
PINK   = 'EC4899'; PINK_L  = 'FCE7F3'
INDIGO = '4F46E5'; IND_L   = 'E0E7FF'
ORANGE = 'EA580C'; ORG_L   = 'FFEDD5'
GRAY   = '64748B'; GRAY_L  = 'F1F5F9'
WHITE  = 'FFFFFF'; DARK    = '1E293B'
LIGHT  = 'F8FAFC'; DIVIDER = 'E2E8F0'

def side(color='DDDDDD'): return Side(style='thin', color=color)
def border(color='DDDDDD'): return Border(left=side(color),right=side(color),top=side(color),bottom=side(color))
def fill(hex): return PatternFill('solid', fgColor=hex)
def font(bold=False,color=DARK,size=10,italic=False): return Font(bold=bold,color=color,name='Arial',size=size,italic=italic)
def align(h='left',v='center',wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

def title_row(ws, row, cols, bg, text, height=36):
    ws.row_dimensions[row].height = height
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=cols)
    c = ws.cell(row=row,column=1,value=text)
    c.font      = font(True,WHITE,14)
    c.fill      = fill(bg)
    c.alignment = align('center')

def subtitle_row(ws, row, cols, bg, text, height=20):
    ws.row_dimensions[row].height = height
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=cols)
    c = ws.cell(row=row,column=1,value=text)
    c.font      = font(False,WHITE,9,True)
    c.fill      = fill(bg)
    c.alignment = align('center')

def header_row(ws, row, headers, bg, height=26):
    ws.row_dimensions[row].height = height
    for i,h in enumerate(headers,1):
        c = ws.cell(row=row,column=i,value=h)
        c.font      = font(True,WHITE,10)
        c.fill      = fill(bg)
        c.alignment = align('center')
        c.border    = border()

def data_cell(ws, row, col, value, bold=False, color=DARK, bg=WHITE, halign='left', num_fmt=None):
    c = ws.cell(row=row,column=col,value=value)
    c.font      = font(bold,color)
    c.fill      = fill(bg)
    c.alignment = align(halign)
    c.border    = border(DIVIDER)
    if num_fmt: c.number_format = num_fmt
    return c

def total_row(ws, row, cols, bg, values, height=26):
    ws.row_dimensions[row].height = height
    for i in range(1,cols+1):
        v = values.get(i,'')
        c = ws.cell(row=row,column=i,value=v)
        c.font      = font(True,WHITE,11)
        c.fill      = fill(bg)
        c.alignment = align('right' if isinstance(v,(int,float)) or (isinstance(v,str) and v.startswith('=')) else 'left')
        c.border    = border()
        if isinstance(v,(int,float)): c.number_format = '#,##0'

def note_row(ws, row, cols, text):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=cols)
    c = ws.cell(row=row,column=1,value=f'💡 {text}')
    c.font      = font(False,AMBER,9,True)
    c.fill      = fill(AMB_L)
    c.alignment = align('left')

def set_widths(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

def row_bg(i): return LIGHT if i%2!=0 else WHITE

def num(v): return float(v or 0)

# ── LOAD DATA ─────────────────────────────────────────────
data      = json.loads(sys.stdin.read())
biz       = data.get('biz','My Business')
date_str  = data.get('date','')
summary   = data.get('summary',{})
monthly   = data.get('monthly',[])
clients   = data.get('clients',[])
suppliers = data.get('suppliers',[])
projects  = data.get('projects',[])
income    = data.get('income',[])
expenses  = data.get('expenses',[])
banking   = data.get('banking',[])
ledger    = data.get('ledger',[])
loans     = data.get('loans',[])
savings   = data.get('savings',[])
assets    = data.get('assets',[])
MONTHS    = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════
# 1. SUMMARY SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('📋 Summary')
set_widths(ws,[32,22,22])
title_row(ws,1,3,BLUE,f'{biz} — Financial Summary')
subtitle_row(ws,2,3,BLUE,f'Generated: {date_str}  |  All amounts in KSh')
ws.append([])

header_row(ws,4,['METRIC','AMOUNT (KSh)','NOTES'],BLUE)
rows_data = [
    ('Total Income',     num(summary.get('totalIncome')),    'All income recorded'),
    ('Total Expenses',   num(summary.get('totalExpenses')),  'All expenses recorded'),
    ('Gross Profit',     num(summary.get('grossProfit')),    '=Income - Expenses'),
    ('They Owe Me',      num(summary.get('totalDebtors')),   'Outstanding client balances'),
    ('I Owe',            num(summary.get('totalCreditors')), 'Outstanding supplier balances'),
    ('Net Position',     num(summary.get('totalDebtors'))-num(summary.get('totalCreditors')), 'Debtors - Creditors'),
]
for i,(label,val,note) in enumerate(rows_data):
    r  = 5+i
    bg = row_bg(i)
    color = GREEN if val>=0 else RED
    ws.row_dimensions[r].height = 24
    data_cell(ws,r,1,label,True,DARK,bg)
    data_cell(ws,r,2,val,True,color,bg,'right','#,##0')
    data_cell(ws,r,3,note,False,GRAY,bg)

# Profit margin formula
r = 11
ws.row_dimensions[r].height = 24
data_cell(ws,r,1,'Profit Margin',True,DARK,LIGHT)
c = ws.cell(row=r,column=2,value='=IF(B5=0,0,B7/B5)')
c.font=font(True,BLUE); c.fill=fill(LIGHT); c.alignment=align('right'); c.border=border(DIVIDER); c.number_format='0.0%'
data_cell(ws,r,3,'Auto-calculated',False,GRAY,LIGHT)

r = 12
ws.row_dimensions[r].height = 24
data_cell(ws,r,1,'Expense Ratio',True,DARK,WHITE)
c = ws.cell(row=r,column=2,value='=IF(B5=0,0,B6/B5)')
c.font=font(True,PURPLE); c.fill=fill(WHITE); c.alignment=align('right'); c.border=border(DIVIDER); c.number_format='0.0%'
data_cell(ws,r,3,'Expenses / Income',False,GRAY,WHITE)

note_row(ws,14,3,'Read-only summary sheet. Edit data in individual sheets below and import back to update the website.')

# ══════════════════════════════════════════════════════════
# 2. MONTHLY P&L SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('📅 Monthly P&L')
set_widths(ws,[12,18,18,18,12,18])
title_row(ws,1,6,BLUE,f'{biz} — Monthly Profit & Loss')
subtitle_row(ws,2,6,BLUE,f'Financial Year {date_str[-4:] if len(date_str)>=4 else "2026"}  |  Editable')
ws.append([])
header_row(ws,4,['MONTH','INCOME','EXPENSES','PROFIT','MARGIN','RUNNING TOTAL'],BLUE)
ds = 5
running = 0
for i,m in enumerate(monthly):
    r   = ds+i
    inc = num(m.get('income'))
    exp = num(m.get('expenses'))
    bg  = row_bg(i)
    ws.row_dimensions[r].height = 22
    data_cell(ws,r,1,MONTHS[i] if i<12 else '',True,DARK,bg)
    data_cell(ws,r,2,inc,False,GREEN,bg,'right','#,##0')
    data_cell(ws,r,3,exp,False,RED,bg,'right','#,##0')
    data_cell(ws,r,4,f'=B{r}-C{r}',True,DARK,bg,'right','#,##0')
    c=ws.cell(row=r,column=5,value=f'=IF(B{r}=0,0,D{r}/B{r})')
    c.font=font(False,BLUE); c.fill=fill(bg); c.alignment=align('right'); c.border=border(DIVIDER); c.number_format='0.0%'
    data_cell(ws,r,6,f'=SUM(D{ds}:D{r})',False,TEAL,bg,'right','#,##0')

tr = ds+len(monthly)
total_row(ws,tr,6,BLUE,{
    1:'TOTAL',
    2:f'=SUM(B{ds}:B{tr-1})',
    3:f'=SUM(C{ds}:C{tr-1})',
    4:f'=SUM(D{ds}:D{tr-1})',
})
note_row(ws,tr+2,6,'Edit Income (col B) and Expenses (col C) values. Profit, Margin and Running Total recalculate automatically.')

# ══════════════════════════════════════════════════════════
# 3. CLIENTS SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('👤 Clients')
set_widths(ws,[24,15,28,15,18,18,16])
title_row(ws,1,7,TEAL,f'{biz} — Client Directory')
subtitle_row(ws,2,7,TEAL,'Editable — Changes will update client records when imported')
ws.append([])
header_row(ws,4,['CLIENT NAME','PHONE','EMAIL','LOCATION','TOTAL BILLED','TOTAL PAID','BALANCE'],TEAL)
ds=5
for i,c in enumerate(clients):
    r   = ds+i
    bil = num(c.get('total_billed'))
    pai = num(c.get('total_paid'))
    bg  = row_bg(i)
    ws.row_dimensions[r].height = 22
    data_cell(ws,r,1,c.get('name',''),True,DARK,bg)
    data_cell(ws,r,2,c.get('phone',''),False,GRAY,bg)
    data_cell(ws,r,3,c.get('email',''),False,GRAY,bg)
    data_cell(ws,r,4,c.get('location',''),False,GRAY,bg)
    data_cell(ws,r,5,bil,False,BLUE,bg,'right','#,##0')
    data_cell(ws,r,6,pai,False,GREEN,bg,'right','#,##0')
    data_cell(ws,r,7,f'=E{r}-F{r}',True,RED if bil>pai else GREEN,bg,'right','#,##0')
if clients:
    tr=ds+len(clients)
    total_row(ws,tr,7,TEAL,{1:'TOTAL',5:f'=SUM(E{ds}:E{tr-1})',6:f'=SUM(F{ds}:F{tr-1})',7:f'=SUM(G{ds}:G{tr-1})'})
note_row(ws,ds+len(clients)+2,7,'Edit name, phone, email, location. Balance auto-calculates from Billed - Paid.')

# ══════════════════════════════════════════════════════════
# 4. DEBTORS SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('📥 Debtors')
set_widths(ws,[24,15,28,15,18,18,16])
title_row(ws,1,7,AMBER,f'{biz} — Debtors Report (Who Owes Me)')
subtitle_row(ws,2,7,AMBER,'Only clients with outstanding balances')
ws.append([])
header_row(ws,4,['CLIENT NAME','PHONE','EMAIL','LOCATION','TOTAL BILLED','TOTAL PAID','BALANCE OWED'],AMBER)
debtors = [c for c in clients if num(c.get('total_billed'))>num(c.get('total_paid'))]
ds=5
for i,c in enumerate(debtors):
    r   = ds+i
    bil = num(c.get('total_billed'))
    pai = num(c.get('total_paid'))
    bg  = row_bg(i)
    ws.row_dimensions[r].height = 22
    data_cell(ws,r,1,c.get('name',''),True,DARK,bg)
    data_cell(ws,r,2,c.get('phone',''),False,GRAY,bg)
    data_cell(ws,r,3,c.get('email',''),False,GRAY,bg)
    data_cell(ws,r,4,c.get('location',''),False,GRAY,bg)
    data_cell(ws,r,5,bil,False,BLUE,bg,'right','#,##0')
    data_cell(ws,r,6,pai,False,GREEN,bg,'right','#,##0')
    data_cell(ws,r,7,f'=E{r}-F{r}',True,RED,bg,'right','#,##0')
if debtors:
    tr=ds+len(debtors)
    total_row(ws,tr,7,AMBER,{1:'TOTAL OWED',5:f'=SUM(E{ds}:E{tr-1})',6:f'=SUM(F{ds}:F{tr-1})',7:f'=SUM(G{ds}:G{tr-1})'})
note_row(ws,ds+len(debtors)+2,7,'Update Total Paid (col F) when a client pays. Balance recalculates automatically.')

# ══════════════════════════════════════════════════════════
# 5. SUPPLIERS SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('🏭 Suppliers')
set_widths(ws,[24,15,28,26,18,18,16])
title_row(ws,1,7,PURPLE,f'{biz} — Supplier Directory')
subtitle_row(ws,2,7,PURPLE,'Editable — Changes will update supplier records when imported')
ws.append([])
header_row(ws,4,['SUPPLIER NAME','PHONE','EMAIL','ITEMS SUPPLIED','TOTAL OWED','TOTAL PAID','BALANCE'],PURPLE)
ds=5
for i,s in enumerate(suppliers):
    r   = ds+i
    owe = num(s.get('total_owed'))
    pai = num(s.get('total_paid'))
    bg  = row_bg(i)
    ws.row_dimensions[r].height = 22
    data_cell(ws,r,1,s.get('name',''),True,DARK,bg)
    data_cell(ws,r,2,s.get('phone',''),False,GRAY,bg)
    data_cell(ws,r,3,s.get('email',''),False,GRAY,bg)
    data_cell(ws,r,4,s.get('items_supplied',''),False,GRAY,bg)
    data_cell(ws,r,5,owe,False,RED,bg,'right','#,##0')
    data_cell(ws,r,6,pai,False,GREEN,bg,'right','#,##0')
    data_cell(ws,r,7,f'=E{r}-F{r}',True,PURPLE,bg,'right','#,##0')
if suppliers:
    tr=ds+len(suppliers)
    total_row(ws,tr,7,PURPLE,{1:'TOTAL',5:f'=SUM(E{ds}:E{tr-1})',6:f'=SUM(F{ds}:F{tr-1})',7:f'=SUM(G{ds}:G{tr-1})'})
note_row(ws,ds+len(suppliers)+2,7,'Edit supplier details. Update Total Paid when you pay a supplier.')

# ══════════════════════════════════════════════════════════
# 6. CREDITORS SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('📤 Creditors')
set_widths(ws,[24,15,28,26,18,18,16])
title_row(ws,1,7,RED,f'{biz} — Creditors Report (Who I Owe)')
subtitle_row(ws,2,7,RED,'Only suppliers with outstanding balances')
ws.append([])
header_row(ws,4,['SUPPLIER NAME','PHONE','EMAIL','ITEMS SUPPLIED','TOTAL OWED','TOTAL PAID','BALANCE'],RED)
creditors = [s for s in suppliers if num(s.get('total_owed'))>num(s.get('total_paid'))]
ds=5
for i,s in enumerate(creditors):
    r   = ds+i
    owe = num(s.get('total_owed'))
    pai = num(s.get('total_paid'))
    bg  = row_bg(i)
    ws.row_dimensions[r].height = 22
    data_cell(ws,r,1,s.get('name',''),True,DARK,bg)
    data_cell(ws,r,2,s.get('phone',''),False,GRAY,bg)
    data_cell(ws,r,3,s.get('email',''),False,GRAY,bg)
    data_cell(ws,r,4,s.get('items_supplied',''),False,GRAY,bg)
    data_cell(ws,r,5,owe,False,RED,bg,'right','#,##0')
    data_cell(ws,r,6,pai,False,GREEN,bg,'right','#,##0')
    data_cell(ws,r,7,f'=E{r}-F{r}',True,RED,bg,'right','#,##0')
if creditors:
    tr=ds+len(creditors)
    total_row(ws,tr,7,RED,{1:'TOTAL OWED',5:f'=SUM(E{ds}:E{tr-1})',6:f'=SUM(F{ds}:F{tr-1})',7:f'=SUM(G{ds}:G{tr-1})'})
note_row(ws,ds+len(creditors)+2,7,'Update Total Paid (col F) when you pay a supplier. Balance recalculates automatically.')

# ══════════════════════════════════════════════════════════
# 7. PROJECTS SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('🔨 Projects')
set_widths(ws,[26,14,18,18,18,16,14,26])
title_row(ws,1,8,INDIGO,f'{biz} — Project Register')
subtitle_row(ws,2,8,INDIGO,'All projects with status and financial details')
ws.append([])
header_row(ws,4,['PROJECT NAME','STATUS','CONTRACT AMT','AMOUNT PAID','BALANCE','COMPLETION','PRODUCT TYPE','NOTES'],INDIGO)
ds=5
for i,p in enumerate(projects):
    r      = ds+i
    camt   = num(p.get('contract_amount'))
    paid   = num(p.get('amount_paid',0))
    status = p.get('status','active')
    s_color = GREEN if status=='completed' else (AMBER if status=='active' else RED)
    bg = row_bg(i)
    ws.row_dimensions[r].height = 22
    data_cell(ws,r,1,p.get('name',''),True,DARK,bg)
    data_cell(ws,r,2,status.upper(),True,s_color,bg,'center')
    data_cell(ws,r,3,camt,False,BLUE,bg,'right','#,##0')
    data_cell(ws,r,4,paid,False,GREEN,bg,'right','#,##0')
    data_cell(ws,r,5,f'=C{r}-D{r}',True,INDIGO,bg,'right','#,##0')
    data_cell(ws,r,6,p.get('estimated_completion',''),False,GRAY,bg,'center')
    data_cell(ws,r,7,p.get('product_type',''),False,GRAY,bg)
    data_cell(ws,r,8,p.get('notes',''),False,GRAY,bg)
if projects:
    tr=ds+len(projects)
    total_row(ws,tr,8,INDIGO,{1:'TOTAL',3:f'=SUM(C{ds}:C{tr-1})',4:f'=SUM(D{ds}:D{tr-1})',5:f'=SUM(E{ds}:E{tr-1})'})
note_row(ws,ds+len(projects)+2,8,'Update Amount Paid (col D) and Status (col B) as projects progress.')

# ══════════════════════════════════════════════════════════
# 8. INCOME ENTRIES SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('💰 Income')
set_widths(ws,[12,18,20,20,22,22,30])
title_row(ws,1,7,GREEN,f'{biz} — Income Tracker')
subtitle_row(ws,2,7,GREEN,'All income entries — Add new rows at the bottom')
ws.append([])
header_row(ws,4,['DATE','AMOUNT (KSh)','COLLECTION POINT','REFERENCE CODE','CLIENT','PROJECT','NOTES'],GREEN)
ds=5
for i,e in enumerate(income):
    r  = ds+i
    bg = row_bg(i)
    ws.row_dimensions[r].height = 22
    data_cell(ws,r,1,e.get('date',''),False,GRAY,bg,'center')
    data_cell(ws,r,2,num(e.get('amount')),True,GREEN,bg,'right','#,##0')
    data_cell(ws,r,3,e.get('collection_point',''),False,BLUE,bg,'center')
    data_cell(ws,r,4,e.get('reference_code',''),False,GRAY,bg,'center')
    data_cell(ws,r,5,e.get('client',''),False,DARK,bg)
    data_cell(ws,r,6,e.get('project',''),False,DARK,bg)
    data_cell(ws,r,7,e.get('notes',''),False,GRAY,bg)
if income:
    tr=ds+len(income)
    total_row(ws,tr,7,GREEN,{1:'TOTAL',2:f'=SUM(B{ds}:B{tr-1})'})
note_row(ws,ds+len(income)+2,7,'Add new income rows below the last entry. Import file to save to website.')

# ══════════════════════════════════════════════════════════
# 9. EXPENSES SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('💸 Expenses')
set_widths(ws,[12,18,16,18,18,22,22,28])
title_row(ws,1,8,RED,f'{biz} — Expense Tracker')
subtitle_row(ws,2,8,RED,'All expense entries — Add new rows at the bottom')
ws.append([])
header_row(ws,4,['DATE','AMOUNT (KSh)','TYPE','PAYMENT METHOD','REFERENCE','SUPPLIER','PROJECT','NOTES'],RED)
ds=5
for i,e in enumerate(expenses):
    r  = ds+i
    bg = row_bg(i)
    ws.row_dimensions[r].height = 22
    data_cell(ws,r,1,e.get('date',''),False,GRAY,bg,'center')
    data_cell(ws,r,2,num(e.get('amount')),True,RED,bg,'right','#,##0')
    data_cell(ws,r,3,e.get('type',''),False,AMBER,bg,'center')
    data_cell(ws,r,4,e.get('collection_point',''),False,BLUE,bg,'center')
    data_cell(ws,r,5,e.get('reference_code',''),False,GRAY,bg,'center')
    data_cell(ws,r,6,e.get('supplier',''),False,DARK,bg)
    data_cell(ws,r,7,e.get('project',''),False,DARK,bg)
    data_cell(ws,r,8,e.get('notes',''),False,GRAY,bg)
if expenses:
    tr=ds+len(expenses)
    total_row(ws,tr,8,RED,{1:'TOTAL',2:f'=SUM(B{ds}:B{tr-1})'})
note_row(ws,ds+len(expenses)+2,8,'Add new expense rows. Type options: materials, salary, transport, utility, general, direct, personal.')

# ══════════════════════════════════════════════════════════
# 10. BANKING ACCOUNTS SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('🏦 Banking')
set_widths(ws,[24,16,18,18,18])
title_row(ws,1,5,TEAL,f'{biz} — Banking Accounts')
subtitle_row(ws,2,5,TEAL,'All payment accounts and balances')
ws.append([])
header_row(ws,4,['ACCOUNT NAME','TYPE','MONEY IN','MONEY OUT','BALANCE'],TEAL)
ds=5
for i,b in enumerate(banking):
    r    = ds+i
    bg   = row_bg(i)
    tin  = num(b.get('total_in'))
    tout = num(b.get('total_out'))
    ws.row_dimensions[r].height = 22
    data_cell(ws,r,1,b.get('name',''),True,DARK,bg)
    data_cell(ws,r,2,b.get('type','').upper(),False,TEAL,bg,'center')
    data_cell(ws,r,3,tin, False,GREEN,bg,'right','#,##0')
    data_cell(ws,r,4,tout,False,RED,  bg,'right','#,##0')
    data_cell(ws,r,5,f'=C{r}-D{r}',True,TEAL,bg,'right','#,##0')
if banking:
    tr=ds+len(banking)
    total_row(ws,tr,5,TEAL,{1:'TOTAL',3:f'=SUM(C{ds}:C{tr-1})',4:f'=SUM(D{ds}:D{tr-1})',5:f'=SUM(E{ds}:E{tr-1})'})
note_row(ws,ds+len(banking)+2,5,'View account balances. Edit account names in col A to update.')

# ══════════════════════════════════════════════════════════
# 11. BANK LEDGER SHEET
# ══════════════════════════════════════════════════════════
if ledger:
    ws = wb.create_sheet('📒 Ledger')
    set_widths(ws,[24,14,22,20,16,16,16])
    title_row(ws,1,7,TEAL,f'{biz} — Full Bank Ledger')
    subtitle_row(ws,2,7,TEAL,'All transactions across all accounts')
    ws.append([])
    header_row(ws,4,['ACCOUNT','DATE','DESCRIPTION','REFERENCE','MONEY IN','MONEY OUT','BALANCE'],TEAL)
    ds=5
    for i,e in enumerate(ledger):
        r  = ds+i
        bg = row_bg(i)
        ws.row_dimensions[r].height = 22
        data_cell(ws,r,1,e.get('account',''),True,DARK,bg)
        data_cell(ws,r,2,e.get('date',''),False,GRAY,bg,'center')
        data_cell(ws,r,3,e.get('description',''),False,DARK,bg)
        data_cell(ws,r,4,e.get('reference',''),False,GRAY,bg,'center')
        data_cell(ws,r,5,num(e.get('amount_in')), False,GREEN,bg,'right','#,##0')
        data_cell(ws,r,6,num(e.get('amount_out')),False,RED,  bg,'right','#,##0')
        data_cell(ws,r,7,num(e.get('balance')),   True, TEAL, bg,'right','#,##0')
    tr=ds+len(ledger)
    total_row(ws,tr,7,TEAL,{1:'TOTAL',5:f'=SUM(E{ds}:E{tr-1})',6:f'=SUM(F{ds}:F{tr-1})'})

# ══════════════════════════════════════════════════════════
# 12. LOANS SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('🏛️ Loans')
set_widths(ws,[26,16,18,18,18,14,18,14])
title_row(ws,1,8,ORANGE,f'{biz} — Loan Manager')
subtitle_row(ws,2,8,ORANGE,'Track all loans borrowed and lent')
ws.append([])
header_row(ws,4,['LOAN NAME / LENDER','TYPE','PRINCIPAL','AMOUNT PAID','BALANCE','INTEREST %','DUE DATE','STATUS'],ORANGE)
ds=5
if loans:
    for i,l in enumerate(loans):
        r      = ds+i
        prin   = num(l.get('principal_amount') or l.get('amount'))
        paid   = num(l.get('amount_paid',0))
        status = l.get('status','active')
        s_col  = GREEN if status=='paid' else (AMBER if status=='active' else RED)
        bg = row_bg(i)
        ws.row_dimensions[r].height = 22
        data_cell(ws,r,1,l.get('name') or l.get('lender_name',''),True,DARK,bg)
        data_cell(ws,r,2,l.get('type','borrowed').upper(),False,ORANGE,bg,'center')
        data_cell(ws,r,3,prin,False,BLUE,bg,'right','#,##0')
        data_cell(ws,r,4,paid,False,GREEN,bg,'right','#,##0')
        data_cell(ws,r,5,f'=C{r}-D{r}',True,RED,bg,'right','#,##0')
        data_cell(ws,r,6,num(l.get('interest_rate',0)),False,AMBER,bg,'right','0.0%')
        data_cell(ws,r,7,l.get('due_date','') or l.get('end_date',''),False,GRAY,bg,'center')
        data_cell(ws,r,8,status.upper(),True,s_col,bg,'center')
    tr=ds+len(loans)
    total_row(ws,tr,8,ORANGE,{1:'TOTAL',3:f'=SUM(C{ds}:C{tr-1})',4:f'=SUM(D{ds}:D{tr-1})',5:f'=SUM(E{ds}:E{tr-1})'})
else:
    ws.row_dimensions[ds].height = 22
    ws.merge_cells(f'A{ds}:H{ds}')
    c=ws.cell(row=ds,column=1,value='No loans recorded yet. Add loans from the website.')
    c.font=font(False,GRAY,10,True); c.fill=fill(LIGHT); c.alignment=align('center')
note_row(ws,ds+max(len(loans),1)+2,8,'Update Amount Paid (col D) as you repay loans. Balance auto-calculates.')

# ══════════════════════════════════════════════════════════
# 13. SAVINGS SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('💵 Savings')
set_widths(ws,[26,18,18,18,18,14,20])
title_row(ws,1,7,GREEN,f'{biz} — Savings Tracker')
subtitle_row(ws,2,7,GREEN,'Track savings goals and progress')
ws.append([])
header_row(ws,4,['SAVINGS GOAL','TARGET AMOUNT','SAVED SO FAR','REMAINING','PROGRESS %','DUE DATE','NOTES'],GREEN)
ds=5
if savings:
    for i,s in enumerate(savings):
        r      = ds+i
        target = num(s.get('target_amount') or s.get('goal_amount'))
        saved  = num(s.get('current_amount') or s.get('saved_amount',0))
        bg = row_bg(i)
        ws.row_dimensions[r].height = 22
        data_cell(ws,r,1,s.get('name') or s.get('goal_name',''),True,DARK,bg)
        data_cell(ws,r,2,target,False,BLUE,bg,'right','#,##0')
        data_cell(ws,r,3,saved, False,GREEN,bg,'right','#,##0')
        data_cell(ws,r,4,f'=B{r}-C{r}',True,RED,bg,'right','#,##0')
        c=ws.cell(row=r,column=5,value=f'=IF(B{r}=0,0,C{r}/B{r})')
        c.font=font(True,TEAL); c.fill=fill(bg); c.alignment=align('right'); c.border=border(DIVIDER); c.number_format='0.0%'
        data_cell(ws,r,6,s.get('target_date','') or s.get('due_date',''),False,GRAY,bg,'center')
        data_cell(ws,r,7,s.get('notes','') or s.get('description',''),False,GRAY,bg)
    tr=ds+len(savings)
    total_row(ws,tr,7,GREEN,{1:'TOTAL',2:f'=SUM(B{ds}:B{tr-1})',3:f'=SUM(C{ds}:C{tr-1})',4:f'=SUM(D{ds}:D{tr-1})'})
else:
    ws.row_dimensions[ds].height = 22
    ws.merge_cells(f'A{ds}:G{ds}')
    c=ws.cell(row=ds,column=1,value='No savings goals recorded yet. Add from the website.')
    c.font=font(False,GRAY,10,True); c.fill=fill(LIGHT); c.alignment=align('center')
note_row(ws,ds+max(len(savings),1)+2,7,'Update Saved So Far (col C). Progress % auto-calculates.')

# ══════════════════════════════════════════════════════════
# 14. ASSETS SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('🏗️ Assets')
set_widths(ws,[26,16,18,16,18,16,14,22])
title_row(ws,1,8,GRAY,f'{biz} — Assets Register')
subtitle_row(ws,2,8,GRAY,'Track all business assets and their current value')
ws.append([])
header_row(ws,4,['ASSET NAME','CATEGORY','PURCHASE VALUE','CURRENT VALUE','DEPRECIATION','PURCHASE DATE','CONDITION','NOTES'],GRAY)
ds=5
if assets:
    for i,a in enumerate(assets):
        r     = ds+i
        pval  = num(a.get('purchase_value') or a.get('cost') or a.get('value'))
        cval  = num(a.get('current_value') or a.get('book_value') or pval)
        depr  = pval - cval
        bg = row_bg(i)
        ws.row_dimensions[r].height = 22
        data_cell(ws,r,1,a.get('name',''),True,DARK,bg)
        data_cell(ws,r,2,a.get('category','') or a.get('type',''),False,GRAY,bg,'center')
        data_cell(ws,r,3,pval,False,BLUE, bg,'right','#,##0')
        data_cell(ws,r,4,cval,False,GREEN,bg,'right','#,##0')
        data_cell(ws,r,5,f'=C{r}-D{r}',True,RED,bg,'right','#,##0')
        data_cell(ws,r,6,a.get('purchase_date','') or a.get('date_acquired',''),False,GRAY,bg,'center')
        data_cell(ws,r,7,a.get('condition','') or a.get('status','good'),False,GRAY,bg,'center')
        data_cell(ws,r,8,a.get('notes','') or a.get('description',''),False,GRAY,bg)
    tr=ds+len(assets)
    total_row(ws,tr,8,GRAY,{1:'TOTAL',3:f'=SUM(C{ds}:C{tr-1})',4:f'=SUM(D{ds}:D{tr-1})',5:f'=SUM(E{ds}:E{tr-1})'})
else:
    ws.row_dimensions[ds].height = 22
    ws.merge_cells(f'A{ds}:H{ds}')
    c=ws.cell(row=ds,column=1,value='No assets recorded yet. Add from the website.')
    c.font=font(False,GRAY,10,True); c.fill=fill(LIGHT); c.alignment=align('center')
note_row(ws,ds+max(len(assets),1)+2,8,'Update Current Value (col D) periodically. Depreciation auto-calculates.')

# ══════════════════════════════════════════════════════════
# 15. BALANCE SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('⚖️ Balance Sheet')
set_widths(ws,[32,22,22])
title_row(ws,1,3,DARK,f'{biz} — Balance Sheet')
subtitle_row(ws,2,3,'374151',f'As at {date_str}  |  Auto-calculated')
ws.append([])

sections = [
    ('ASSETS',    TEAL,  [('Cash & Bank Balances',sum(num(b.get('balance',0)) for b in banking)),
                           ('Debtors (They Owe Me)',num(summary.get('totalDebtors',0))),
                           ('Projects (Contract Value)',sum(num(p.get('contract_amount',0)) for p in projects)),
                           ('Fixed Assets',sum(num(a.get('current_value') or a.get('purchase_value') or a.get('value',0)) for a in assets))]),
    ('LIABILITIES',RED,  [('Creditors (I Owe)',num(summary.get('totalCreditors',0))),
                           ('Loans Outstanding',sum(num(l.get('principal_amount') or l.get('amount',0))-num(l.get('amount_paid',0)) for l in loans))]),
]
current_row = 4
for section_name, color, items in sections:
    header_row(ws,current_row,[section_name,'AMOUNT (KSh)',''],color)
    current_row += 1
    section_start = current_row
    for j,(label,val) in enumerate(items):
        bg = row_bg(j)
        ws.row_dimensions[current_row].height = 22
        data_cell(ws,current_row,1,f'  {label}',False,DARK,bg)
        data_cell(ws,current_row,2,val,True,BLUE if section_name=='ASSETS' else RED,bg,'right','#,##0')
        data_cell(ws,current_row,3,'',False,DARK,bg)
        current_row += 1
    total_row(ws,current_row,3,color,{1:f'TOTAL {section_name}',2:f'=SUM(B{section_start}:B{current_row-1})'})
    current_row += 2

note_row(ws,current_row,3,'Balance sheet auto-calculates from your data. Keep records updated for accuracy.')

# ══════════════════════════════════════════════════════════
# 16. CASH FLOW SHEET
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('💧 Cash Flow')
set_widths(ws,[28,18,18,18,18])
title_row(ws,1,5,BLUE,f'{biz} — Cash Flow Statement')
subtitle_row(ws,2,5,BLUE,f'Monthly cash movements  |  {date_str[-4:] if len(date_str)>=4 else "2026"}')
ws.append([])
header_row(ws,4,['MONTH','CASH IN','CASH OUT','NET FLOW','CUMULATIVE'],BLUE)
ds=5
cumulative = 0
for i,m in enumerate(monthly):
    r   = ds+i
    inc = num(m.get('income'))
    exp = num(m.get('expenses'))
    bg  = row_bg(i)
    net = inc - exp
    ws.row_dimensions[r].height = 22
    data_cell(ws,r,1,MONTHS[i] if i<12 else '',True,DARK,bg)
    data_cell(ws,r,2,inc,False,GREEN,bg,'right','#,##0')
    data_cell(ws,r,3,exp,False,RED,  bg,'right','#,##0')
    data_cell(ws,r,4,f'=B{r}-C{r}',True,BLUE if net>=0 else RED,bg,'right','#,##0')
    data_cell(ws,r,5,f'=SUM(D{ds}:D{r})',False,TEAL,bg,'right','#,##0')
tr=ds+len(monthly)
total_row(ws,tr,5,BLUE,{1:'TOTAL',2:f'=SUM(B{ds}:B{tr-1})',3:f'=SUM(C{ds}:C{tr-1})',4:f'=SUM(D{ds}:D{tr-1})'})
note_row(ws,tr+2,5,'Cash flow tracks money in vs out each month. Cumulative shows running total.')

# ══════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════
wb.save('/home/claude/full_export.xlsx')
print('SUCCESS')
